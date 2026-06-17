"""
Generate a COMPLETE Feb_summary.xlsx from feb_basic, replicating how jan_summary
is built from jan_basic.

Decoded logic (all verified by reproducing January):
  * jan_summary = the Original + Summary + Services(Sheet3) sheets of the basic
    workbook. Summary & Services are FORMULA sheets that reference Original, so once
    Original holds the target month's numbers, Excel recomputes the dashboard.
  * Every Original block maps to a Percentages section by header (normalising a
    leading "#." and whitespace). Three block shapes occur:
      - SUBQUESTION-TRANSPOSE  (data in Original cols F:Base, G..K:options; rows =
        sub-questions): A2/A4A/A6/A8/A10 (Website/Mobile), B1 (Call Center/CSC),
        B3 (Call/Service centre), D1 (CSC), C3 (Email), E2 (Complaint).
        Source: section's sub-question COLUMNS (positional) x option ROWS.
      - CHANNEL-HAPPINESS-TRANSPOSE (Original rows = channels, cols F..K options):
        A12 (Web/Mob), B5 (Call/CSC), C5 (Email). Source: the section's option ROWS
        x the global channel banner COLUMN (matched by label -> Jan/Feb col shift).
      - DIRECT CROSSTAB (Original rows = options, cols B:Total C..G:channels):
        Overall Satisfaction, X14 (Value), Y14 (Trust). Source: option ROWS x
        channel banner COLUMNS (by label) + Total col B.
  * NPS (A16) block: Promoter = mean of scores 9-10, Passive 7-8, Detractor 0-6,
    NPS = Promoter - Detractor, per channel + total (rows 141-146, cols S-V).
  * Computed columns: O = T2B (Strongly Agree + Agree) for transpose blocks; the
    direct blocks carry their own T2B option row.

Usage:
  python build_feb_summary.py --validate              # rebuild Jan, diff vs real
  python build_feb_summary.py --out outputs/Feb_summary.xlsx
"""
from __future__ import annotations
import argparse, re
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

JAN_BASIC = "data/jan_basic/MOHAP CSAT Tabs for Jan-2026.xlsx"
FEB_BASIC = "data/feb_basic/MOHAP CSAT Tabs - Feb 2026.xlsx"
JAN_SUMMARY = "data/jan_summary.xlsx"

OPT5 = ["Strongly Agree", "Agree", "Neutral", "Disagree", "Strongly Disagree"]
HAPPY5 = ["Very Happy", "Happy", "Neutral", "Unhappy", "Very Unhappy"]
CHANNELS = ["Call Center", "Customer service center", "Website", "Mobile app", "Email"]
TOL = 0.0015


def norm(h):
    if not isinstance(h, str):
        return ""
    h = h.strip()
    if h.startswith("#."):
        h = h[2:].strip()
    return re.sub(r"\s+", " ", h)


def grid(path, sheet):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def channel_map(prows):
    for row in prows[:4]:
        cols = {}
        for i, v in enumerate(row):
            if isinstance(v, str) and v.strip() in CHANNELS and v.strip() not in cols:
                cols[v.strip()] = i  # 0-based
        if all(c in cols for c in CHANNELS):
            return cols
    return {}


def find_sections(prows):
    """{norm header: {row, base_row, opt{label:row}, subcols:[0-based]}}"""
    secs = {}
    for r in range(len(prows)):
        a = prows[r][0]
        if not isinstance(a, str) or not a.strip():
            continue
        # a section = header followed (within 2 rows) by a 'Base' row
        base_row = None
        for d in (1, 2):
            if r + d < len(prows):
                lab = prows[r + d][0]
                if isinstance(lab, str) and lab.strip().lower().startswith("base"):
                    base_row = r + d
                    break
        if base_row is None:
            continue
        opt = {}
        rr = base_row + 1
        while rr < len(prows) and rr < base_row + 9:
            lab = prows[rr][0]
            ls = str(lab).strip() if lab is not None else ""
            if ls in OPT5 or ls in HAPPY5:
                opt[ls] = rr
            if ls in ("Total", "____") or (ls == "" and opt):
                break
            rr += 1
        # subcols: contiguous numeric/'present' base cells from col B (idx1)
        subcols = []
        brow = prows[base_row]
        for c in range(1, len(brow)):
            v = brow[c]
            if v is not None and str(v).strip() != "":
                subcols.append(c)
            elif subcols:
                break
        secs.setdefault(norm(a), {"row": r, "base_row": base_row, "opt": opt, "subcols": subcols})
    return secs


def feb_demographic_cols(prows):
    """{banner label -> 0-based col} for the Gender / Nationality / Age /
    Emirate-of-residence banner groups ONLY. Bounded by the row-1 group headers so
    the S4 'Customer service center visited' emirate columns (which repeat the same
    emirate names) are excluded."""
    if len(prows) < 2:
        return {}
    row_groups, row_labels = prows[0], prows[1]
    headers = sorted((i, v.strip()) for i, v in enumerate(row_groups)
                     if isinstance(v, str) and v.strip())
    wanted = ("gender", "nationality", "age gr", "emirate of residence")
    out = {}
    for k, (c, htext) in enumerate(headers):
        if not any(w in htext.lower() for w in wanted):
            continue
        end = headers[k + 1][0] if k + 1 < len(headers) else len(row_groups)
        for cc in range(c, end):
            lab = row_labels[cc] if cc < len(row_labels) else None
            if isinstance(lab, str) and lab.strip():
                out.setdefault(lab.strip(), cc)
    return out


# ── Original block detection ──────────────────────────────────────────────────

def orig_blocks(orows):
    """Return [(norm_header, header_row, [data_rows], kind)].
    Header rows come in three flavours:
      q  question block  -> col E holds the section header, col F starts 'Base'
      d  dashboard cross -> col A is 'Overall Sati' / X14 / Y14
      n  NPS (A16)       -> col C starts 'A16'
    Data rows = every row up to the next header (blanks are harmless)."""
    n = len(orows)
    cell = lambda r, c: orows[r][c] if c < len(orows[r]) else None
    starts = []
    for r in range(n):
        e, f, aa, cc = cell(r, 4), cell(r, 5), cell(r, 0), cell(r, 2)
        if isinstance(e, str) and e.strip() and isinstance(f, str) and f.strip().lower().startswith("base"):
            starts.append((r, norm(e), "q"))
        elif isinstance(aa, str) and ("Overall Sati" in aa or re.match(r"#?\.?\s*[XY]14", aa.strip())):
            starts.append((r, norm(aa), "d"))
        elif isinstance(cc, str) and cc.strip().startswith("A16"):
            starts.append((r, norm(cc), "n"))
    blocks = []
    for i, (r, hdr, kind) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else n
        blocks.append((hdr, r, list(range(r + 1, end)), kind))
    return blocks


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--validate", action="store_true")
    ap.add_argument("--out", default="outputs/Feb_summary.xlsx")
    ap.add_argument("--basic", default=FEB_BASIC)
    a = ap.parse_args()
    basic = JAN_BASIC if a.validate else a.basic

    twb = openpyxl.load_workbook(JAN_SUMMARY, data_only=False)
    o = twb["Original"]
    prows = grid(basic, "Percentages")
    secs = find_sections(prows)
    chmap = channel_map(prows)
    tmpl = grid(JAN_SUMMARY, "Original")          # for layout reference (jan)
    blocks = orig_blocks(tmpl)

    log = []

    def put(rr, cc, val):
        o.cell(rr + 1, cc + 1).value = val       # rr,cc 0-based

    overallsati = secs.get(norm("Overall Satiufcation (A12+B5+C5)"))
    feb_demo = feb_demographic_cols(prows)

    def fill_overall_sat_demographics(sec):
        """Populate the Overall-Satisfaction block's DEMOGRAPHIC banner columns in
        Original (gender/nationality/age/city). The Summary 'A16' block (T33:W55)
        reads base (row 151) and T2B (row 160) from these; direct_block only fills
        Total + channel columns, so without this they stay at January."""
        if sec is None or not feb_demo:
            return 0
        titles = [r for r in range(len(tmpl))
                  if isinstance(tmpl[r][0], str) and "Overall Sati" in tmpl[r][0]]
        for title in titles:
            banner = None
            for br in range(max(0, title - 4), title + 1):
                b = tmpl[br][1] if len(tmpl[br]) > 1 else None
                if isinstance(b, str) and b.strip() == "Total":
                    banner = br
            if banner is None:
                continue
            demo_map = {c: feb_demo[lab.strip()]
                        for c, lab in enumerate(tmpl[banner])
                        if isinstance(lab, str) and lab.strip() in feb_demo
                        and lab.strip() not in CHANNELS and lab.strip() != "Total"}
            if not demo_map:
                continue
            n = 0
            for dr in range(title + 1, min(len(tmpl), title + 15)):
                ol = str(tmpl[dr][0]).strip() if isinstance(tmpl[dr][0], str) else ""
                is_t2b = ol.startswith("T2B")
                src = sec["base_row"] if ol.lower().startswith("base") else sec["opt"].get(ol)
                if src is None and not is_t2b:
                    continue
                for ocol, fcol in demo_map.items():
                    if is_t2b:
                        g = num(prows[sec["opt"]["Very Happy"]][fcol]) if "Very Happy" in sec["opt"] else None
                        h = num(prows[sec["opt"]["Happy"]][fcol]) if "Happy" in sec["opt"] else None
                        val = round(g + (h or 0), 4) if g is not None else None
                    else:
                        val = prows[src][fcol] if fcol < len(prows[src]) else None
                    if val is not None:
                        put(dr, ocol, val)
                n += 1
            return n
        return 0

    def transpose_block(data, sec, labels, by_channel):
        # only rows carrying a sub-question / channel label in col C participate, so
        # positional alignment to the section's sub-columns is exact.
        rows = [dr for dr in data if isinstance(tmpl[dr][2], str) and tmpl[dr][2].strip()]
        for i, dr in enumerate(rows):
            if by_channel:
                pc = chmap.get(tmpl[dr][2].strip())
            else:
                pc = sec["subcols"][i] if i < len(sec["subcols"]) else None
            if pc is None:
                continue
            put(dr, 5, prows[sec["base_row"]][pc])               # F base
            for j, lab in enumerate(labels):
                pr = sec["opt"].get(lab)
                if pr is not None:
                    put(dr, 6 + j, prows[pr][pc])
            g = num(prows[sec["opt"][labels[0]]][pc]) if labels[0] in sec["opt"] else None
            h = num(prows[sec["opt"][labels[1]]][pc]) if labels[1] in sec["opt"] else None
            if g is not None:                              # O = T2B (dash 2nd box = 0)
                put(dr, 14, round(g + (h or 0), 4))

    def direct_block(data, sec, labels):
        # Original rows = labelled options; cols B(Total,idx1) C..G(idx2..6) channels
        def srccol(ci):
            return 1 if ci == 1 else chmap.get(CHANNELS[ci - 2])
        for dr in data:
            ol = str(tmpl[dr][0]).strip() if isinstance(tmpl[dr][0], str) else ""
            prow = None
            if ol.lower().startswith("base"):
                prow = sec["base_row"]
            elif ol in sec["opt"]:
                prow = sec["opt"][ol]
            elif ol.startswith("T2B"):
                for ci in range(1, 7):
                    sc = srccol(ci)
                    if sc is None:
                        continue
                    g = num(prows[sec["opt"][labels[0]]][sc]) if labels[0] in sec["opt"] else None
                    h = num(prows[sec["opt"][labels[1]]][sc]) if labels[1] in sec["opt"] else None
                    if g is not None:
                        put(dr, ci, round(g + (h or 0), 4))
                continue
            if prow is None:
                continue
            for ci in range(1, 7):
                sc = srccol(ci)
                if sc is not None and sc < len(prows[prow]):
                    put(dr, ci, prows[prow][sc])

    # NPS (A16): Promoter = scores 9-10, Passive 7-8, Detractor 0-6, per channel.
    # Original cols S(18)=Promoter, T(19)=Passive, U(20)=Detractor; V=S-U is a formula.
    def nps_block(data):
        a16 = next((r for r in range(len(prows))
                    if isinstance(prows[r][0], str) and prows[r][0].strip().startswith("A16")), None)
        if a16 is None:
            return
        scores = {}
        for rr in range(a16 + 1, min(a16 + 16, len(prows))):
            lab = str(prows[rr][0]).strip() if prows[rr][0] is not None else ""
            m = re.match(r"(\d+)", lab)
            if m:
                scores[int(m.group(1))] = rr
        for dr in data:
            ol = str(tmpl[dr][0]).strip() if isinstance(tmpl[dr][0], str) else ""
            col = 1 if ol == "Total" else chmap.get(ol)
            if col is None:
                continue
            s = lambda ks: sum((num(prows[scores[k]][col]) or 0) for k in ks if k in scores)
            prom, detr = s([9, 10]), s(range(0, 7))
            put(dr, 18, round(prom, 4))            # S Promoter
            put(dr, 19, round(s([7, 8]), 4))       # T Passive
            put(dr, 20, round(detr, 4))            # U Detractor
            put(dr, 21, round(prom - detr, 4))     # V NPS (overwrites =S-U formula)
        log.append(("A16 NPS", "nps", len(data)))

    for hdr, hr, data, kind in blocks:
        low = hdr.lower()
        if kind == "n":
            nps_block(data)
            continue
        # channel-happiness blocks (A12/B5/C5) -> Overall-Sat section by channel
        if "please rate your overall happiness" in low:
            if overallsati:
                transpose_block(data, overallsati, HAPPY5, by_channel=True)
                log.append((hdr[:34], "chan-happiness", len(data)))
            continue
        sec = secs.get(hdr)
        if sec is None:
            continue
        # "Neutral" is shared by both scales -> key on the distinctive "Very Happy"
        labels = HAPPY5 if "Very Happy" in sec["opt"] else OPT5
        if kind == "d":                       # direct dashboard crosstab (Overall/X14/Y14)
            direct_block(data, sec, labels)
            log.append((hdr[:34], "direct-crosstab", len(data)))
        else:                                  # subq-transpose
            transpose_block(data, sec, labels, by_channel=False)
            log.append((hdr[:34], "subq-transpose", len(data)))

    # A16 block source: Overall-Satisfaction T2B by gender/nationality/age/city
    ndemo = fill_overall_sat_demographics(overallsati)
    log.append(("Overall-Sat by demographic (A16 block)", "overall-sat-demo", ndemo))

    # ── validation / save ────────────────────────────────────────────────────
    if a.validate:
        real = grid(JAN_BASIC, "Original")
        import re as _re
        sm = openpyxl.load_workbook(JAN_SUMMARY, data_only=False)["Summary"]
        refs = set()
        for r in range(1, sm.max_row + 1):
            for c in range(1, sm.max_column + 1):
                v = sm.cell(r, c).value
                if isinstance(v, str) and v.startswith("="):
                    for m in _re.finditer(r"Original!\$?([A-Z]{1,3})\$?(\d+)", v):
                        refs.add((column_index_from_string(m.group(1)) - 1, int(m.group(2)) - 1))
        mism = chk = 0
        for c, r in sorted(refs):
            built = o.cell(r + 1, c + 1).value
            realv = real[r][c] if r < len(real) and c < len(real[r]) else None
            if isinstance(realv, str) or realv is None:
                continue  # labels
            if isinstance(built, str) and built.startswith("="):
                continue  # template formula (e.g. NPS V=S-U) recomputes in Excel
            chk += 1
            b = num(built)
            if b is None or abs(b - realv) > TOL:
                mism += 1
                if mism <= 25:
                    print(f"  MISS {get_column_letter(c+1)}{r+1}: built={built} real={realv}")
        print(f"\nblocks processed: {len(log)}")
        for h, k, n in log:
            print(f"   {k:16} {h}")
        print(f"\nSummary-referenced NUMERIC Original cells: checked={chk} mismatches={mism}")
    else:
        import compute_summary
        nc, nt = compute_summary.recompute(twb)   # evaluate Summary formulas -> values
        Path(a.out).parent.mkdir(parents=True, exist_ok=True)
        twb.save(a.out)
        print(f"saved {a.out}  blocks={len(log)}  summary_formulas_computed={nc}/{nt}")


if __name__ == "__main__":
    main()
